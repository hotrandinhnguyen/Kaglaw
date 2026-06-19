"""Excel report builder.

8 sheets:
1. Summary        — Best public/private score per (competition, nick) + # subs + best rank
2. Experiments    — One row per run: params (one col each) + parsed metric, for research
3. Notebooks      — Registered notebooks (id, title, flags, # runs, path)
4. Runs           — Nick + Notebook version + status + runtime
5. Submissions    — Competition + score + LB rank
6. GPU Usage      — Per-nick aggregate runtime (= used hours estimate)
7. Budgets        — Per-nick free slots + remaining GPU hours estimate
8. Outputs & Logs — Per-run output dir + log summary
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import db
from .config import EXPORTS_DIR

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def _runs_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Runs")
    headers = [
        "Run ID", "Nick", "Kaggle Username", "Notebook", "Slug",
        "Version", "Status", "GPU", "TPU", "Pushed At",
        "Completed At", "Runtime (s)", "Runtime (min)", "Error",
    ]
    ws.append(headers)
    with db.connect() as con:
        rows = con.execute(
            """SELECT r.*, n.title AS notebook_title, a.username AS username
               FROM runs r
               LEFT JOIN notebooks n ON n.id=r.notebook_id
               LEFT JOIN accounts a ON a.nick=r.account_nick
               ORDER BY r.id DESC""").fetchall()
    for r in rows:
        rt = r["runtime_seconds"] or 0
        ws.append([
            r["id"],
            r["account_nick"],
            r["username"],
            r["notebook_title"],
            r["slug"],
            r["version_number"],
            r["status"],
            "Y" if r["used_gpu"] else "",
            "Y" if r["used_tpu"] else "",
            r["pushed_at"],
            r["completed_at"],
            round(rt, 1) if rt else None,
            round(rt / 60, 2) if rt else None,
            (r["error"] or "")[:500],
        ])
    _style_header(ws, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"


def _submissions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Submissions")
    headers = [
        "Sub ID", "Competition", "Nick", "File", "Description",
        "Submitted At", "Public Score", "Private Score", "Status",
        "LB Rank (public)", "LB Size", "Last Synced",
    ]
    ws.append(headers)
    with db.connect() as con:
        rows = con.execute("SELECT * FROM submissions ORDER BY competition, account_nick, id DESC").fetchall()
    for r in rows:
        ws.append([
            r["id"], r["competition"], r["account_nick"], r["file_name"],
            r["description"], r["submitted_at"], r["public_score"], r["private_score"],
            r["status"], r["rank_public"], r["leaderboard_size"], r["last_synced"],
        ])
    _style_header(ws, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"


def _gpu_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("GPU Usage")
    headers = [
        "Nick", "Username", "GPU Runs", "GPU Seconds", "GPU Hours",
        "TPU Runs", "TPU Seconds", "TPU Hours",
        "CPU Runs", "CPU Seconds", "CPU Hours",
        "Total Runs",
    ]
    ws.append(headers)
    with db.connect() as con:
        rows = con.execute(
            """SELECT r.account_nick, a.username, r.used_gpu, r.used_tpu,
                      COALESCE(r.runtime_seconds, 0) AS rt
               FROM runs r LEFT JOIN accounts a ON a.nick=r.account_nick"""
        ).fetchall()

    agg: dict[str, dict[str, float]] = defaultdict(lambda: {
        "username": "", "gpu_n": 0, "gpu_s": 0.0,
        "tpu_n": 0, "tpu_s": 0.0, "cpu_n": 0, "cpu_s": 0.0, "total": 0,
    })
    for r in rows:
        nick = r["account_nick"]
        agg[nick]["username"] = r["username"] or ""
        rt = float(r["rt"] or 0)
        agg[nick]["total"] += 1
        if r["used_gpu"]:
            agg[nick]["gpu_n"] += 1
            agg[nick]["gpu_s"] += rt
        elif r["used_tpu"]:
            agg[nick]["tpu_n"] += 1
            agg[nick]["tpu_s"] += rt
        else:
            agg[nick]["cpu_n"] += 1
            agg[nick]["cpu_s"] += rt

    for nick, a in sorted(agg.items()):
        ws.append([
            nick, a["username"],
            int(a["gpu_n"]), round(a["gpu_s"], 1), round(a["gpu_s"] / 3600, 2),
            int(a["tpu_n"]), round(a["tpu_s"], 1), round(a["tpu_s"] / 3600, 2),
            int(a["cpu_n"]), round(a["cpu_s"], 1), round(a["cpu_s"] / 3600, 2),
            int(a["total"]),
        ])
    _style_header(ws, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"

    # Note row
    ws.cell(row=len(agg) + 3, column=1,
            value="Note: Kaggle does not expose GPU quota via API. Hours = sum of run "
                  "runtimes (completed_at - pushed_at) for runs flagged GPU/TPU. Treat "
                  "as estimate, not exact quota.")


def _notebooks_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Notebooks")
    headers = [
        "ID", "Title", "Language", "Type", "GPU", "TPU", "Internet",
        "# Runs", "Local Path", "Created At",
    ]
    ws.append(headers)
    with db.connect() as con:
        rows = con.execute(
            """SELECT n.*, (SELECT COUNT(*) FROM runs r WHERE r.notebook_id=n.id) AS n_runs
               FROM notebooks n ORDER BY n.id DESC"""
        ).fetchall()
    for r in rows:
        ws.append([
            r["id"], r["title"], r["language"], r["kernel_type"],
            "Y" if r["enable_gpu"] else "", "Y" if r["enable_tpu"] else "",
            "Y" if r["enable_internet"] else "", r["n_runs"],
            r["local_path"], r["created_at"],
        ])
    _style_header(ws, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"


def _summary_sheet(wb: Workbook) -> None:
    """Best public score per (competition, nick) + submission counts — the headline stats."""
    ws = wb.create_sheet("Summary")
    headers = [
        "Competition", "Nick", "Best Public", "Best Private",
        "# Subs", "Best LB Rank", "LB Size", "Last Submit",
    ]
    ws.append(headers)
    with db.connect() as con:
        rows = con.execute(
            """SELECT competition, account_nick,
                      MAX(CAST(public_score AS REAL))  AS best_public,
                      MAX(CAST(private_score AS REAL)) AS best_private,
                      COUNT(*) AS n_subs,
                      MIN(rank_public) AS best_rank,
                      MAX(leaderboard_size) AS lb_size,
                      MAX(submitted_at) AS last_submit
               FROM submissions
               GROUP BY competition, account_nick
               ORDER BY competition, best_public DESC"""
        ).fetchall()
    for r in rows:
        ws.append([
            r["competition"], r["account_nick"],
            r["best_public"], r["best_private"], r["n_subs"],
            r["best_rank"], r["lb_size"], r["last_submit"],
        ])
    _style_header(ws, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"
    if ws.max_row == 1:
        ws.cell(row=2, column=1,
                value="No submissions tracked yet. Use sync_submissions_for_competition first.")


def _experiments_sheet(wb: Workbook) -> None:
    """One row per experiment (run), with a column per param + the parsed metric."""
    from . import actions

    ws = wb.create_sheet("Experiments")
    exps = actions.list_experiments(limit=1000)
    param_keys: list[str] = []
    for e in exps:
        for k in e.get("params", {}):
            if k not in param_keys:
                param_keys.append(k)

    fixed = [
        "Run ID", "Notebook", "Nick", "Competition", "Status",
        "Metric", "Value", "LB Public", "Runtime (min)", "Batch", "Tags",
    ]
    headers = fixed + [f"param: {k}" for k in param_keys] + ["Code Snapshot"]
    ws.append(headers)
    for e in exps:
        rt = e.get("runtime_seconds") or 0
        row = [
            e.get("id"), e.get("notebook_title"), e.get("account_nick"),
            e.get("competition"), e.get("status"),
            e.get("metric_name"), e.get("metric_value"), e.get("lb_public"),
            round(rt / 60, 2) if rt else None,
            e.get("batch_id"), e.get("tags"),
        ]
        row += [e.get("params", {}).get(k) for k in param_keys]
        row += [e.get("code_snapshot_path")]
        ws.append(row)
    _style_header(ws, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"
    if ws.max_row == 1:
        ws.cell(row=2, column=1, value="No experiments yet. Push a notebook to create runs.")


def _outputs_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Outputs & Logs")
    headers = ["Run ID", "Nick", "Slug", "Version", "Status", "Output Path", "Log Summary"]
    ws.append(headers)
    with db.connect() as con:
        rows = con.execute(
            "SELECT id, account_nick, slug, version_number, status, output_path, log_summary "
            "FROM runs WHERE output_path IS NOT NULL OR log_summary IS NOT NULL "
            "ORDER BY id DESC").fetchall()
    for r in rows:
        ws.append([
            r["id"], r["account_nick"], r["slug"], r["version_number"],
            r["status"], r["output_path"], (r["log_summary"] or "")[:1000],
        ])
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).alignment = _WRAP
    _style_header(ws, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"


def _budgets_sheet(wb: Workbook) -> None:
    from . import accounts as account_mod
    from . import budgets

    ws = wb.create_sheet("Budgets")
    headers = ["Nick", "Running Now", "Free Slots", "GPU Hours 7d",
               "GPU Hours Remaining", "Submissions Today"]
    ws.append(headers)
    for u in budgets.all_usages([a.nick for a in account_mod.list_accounts()]):
        ws.append([
            u["nick"], u["running_now"], u["free_slots"],
            u["gpu_hours_7d"], u["gpu_hours_remaining"], u["submissions_today"],
        ])
    _style_header(ws, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"
    ws.cell(row=ws.max_row + 2, column=1,
            value="Estimates only — Kaggle has no quota API. GPU hours = sum of GPU run "
                  "runtimes in the last 7 days; remaining = weekly budget minus that.")


def build_workbook() -> Workbook:
    wb = Workbook()
    # remove the default sheet
    default = wb.active
    wb.remove(default)
    _summary_sheet(wb)
    _experiments_sheet(wb)
    _notebooks_sheet(wb)
    _runs_sheet(wb)
    _submissions_sheet(wb)
    _gpu_sheet(wb)
    _budgets_sheet(wb)
    _outputs_sheet(wb)
    return wb


def export_to_file(path: Path | None = None) -> Path:
    if path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = EXPORTS_DIR / f"kaglaw_{ts}.xlsx"
    wb = build_workbook()
    wb.save(path)
    return path
